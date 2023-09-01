import openpyxl

book=openpyxl.load_workbook("C:\\Users\\hp\\Desktop\\PythonDemo.xlsx")
sheet=book.active
Dict={}

# read data from excel sheet
# read=sheet.cell(row=1, column=2)
# print(read.value)

# # write data in excel sheet
# write=sheet.cell(row=2,column=2).value="Demo"
# print(write)

# print(sheet.max_row)
# print(sheet.max_column)

# print(sheet['A5'].value)


# print 1 column value 
# for i in range (1,sheet.max_row+1):
#     print(sheet.cell(row=i,column=1).value)
    
 # print 1 row value    
# for i in range (1,sheet.max_column+1):
#     print(sheet.cell(row=1,column=i).value)  


# print all excel sheet data
# for i in range (1,sheet.max_row+1):
#     for j in range (1,sheet.max_column+1):
#         print(sheet.cell(row=i,column=j).value)  


# print data complete row data with condition
for i in range (1,sheet.max_row+1):   #to get i=3  ,we can change value 1,2,3,
    if sheet.cell(row=i,column=1).value == "TestCase1":
        for j in range (2,sheet.max_column+1):  #to get column j=4,we can change value 1,2,3,
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
            # print(sheet.cell(row=i,column=j).value)
print(Dict)            
        