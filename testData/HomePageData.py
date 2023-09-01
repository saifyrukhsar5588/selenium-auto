import openpyxl


class homgpagedata:
    
    test_home_page_Data = {"FirstName" : "Amir","email" : "aamir@gmail.com","Gendre" : "Male"},{"FirstName" : "sone","email" : "sonu@gmail.com","Gendre" : "Female"}
    []
    
    @staticmethod
    def gettestdata(test_case_name):
        book=openpyxl.load_workbook("C:\\Users\\hp\\Desktop\\PythonDemo.xlsx")
        sheet=book.active
        Dict={}
        for i in range (1,sheet.max_row+1):   #to get i=3  ,we can change value 1,2,3,
         if sheet.cell(row=i,column=1).value == test_case_name:
          for j in range (2,sheet.max_column+1):  #to get column j=4,we can change value 1,2,3,
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
            # print(sheet.cell(row=i,column=j).value)
        return [Dict]            
        